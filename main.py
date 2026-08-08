import json
import os
import shutil
import sqlite3
import tempfile
from datetime import date
from io import BytesIO
from typing import List

from fastapi import FastAPI, Request, Form, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from database import (
    get_db,
    hash_password,
    init_db,
    get_official_check_in_time,
    format_time_display,
    get_school_open_status,
    is_school_open,
    get_teacher_attendance,
    record_check_in,
    get_progress_stats,
    get_session_progress_totals,
    get_monthly_progress_breakdown,
    get_monthly_trend,
    get_monthly_attendance_time_trend,
    get_all_teachers_average_trend,
    bulk_set_school_days,
    get_school_calendar_events,
    count_total_open_days_marked,
    set_setting,
)

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="mps-secret-key-2026")
templates = Jinja2Templates(directory="templates")


@app.on_event("startup")
def startup():
    init_db()


def current_user(request: Request):
    return request.session.get("user")


def require_login(request: Request):
    user = current_user(request)
    if not user:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    return user


def require_admin(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    return user


def shift_month(year: int, month: int, delta: int):
    m = month + delta
    y = year
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    return y, m


# =================== AUTH & STATIC PAGES ===================
@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    return RedirectResponse(url="/login")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if current_user(request):
        user = current_user(request)
        return RedirectResponse(url="/admin" if user["role"] == "admin" else "/dashboard")
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login", response_class=HTMLResponse)
def login_post(request: Request, email: str = Form(...), password: str = Form(...)):
    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE email = ? AND password = ?",
        (email.strip().lower(), hash_password(password)),
    ).fetchone()
    conn.close()

    if not user:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid email or password."})
    if user["status"] == "pending":
        return templates.TemplateResponse("login.html", {"request": request, "error": "Your account is pending admin approval."})
    if user["status"] == "rejected":
        return templates.TemplateResponse("login.html", {"request": request, "error": "Your account has been rejected."})
    if user["is_active"] == 0:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Your account has been deactivated. Contact the admin."})

    request.session["user"] = {
        "id": user["id"],
        "email": user["email"],
        "full_name": user["full_name"],
        "father_name": user["father_name"],
        "qualifications": user["qualifications"],
        "experience": user["experience"],
        "role": user["role"],
        "status": user["status"],
    }

    if user["role"] == "admin":
        return RedirectResponse(url="/admin", status_code=303)
    return RedirectResponse(url="/dashboard", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/signup", response_class=HTMLResponse)
def signup_page(request: Request):
    return templates.TemplateResponse("signup.html", {"request": request, "error": None, "success": None})


@app.post("/signup", response_class=HTMLResponse)
def signup_post(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(...),
    father_name: str = Form(...),
    qualifications: str = Form(...),
    experience: str = Form(...),
):
    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE email = ?", (email.strip().lower(),)).fetchone()
    if existing:
        conn.close()
        return templates.TemplateResponse("signup.html", {
            "request": request, "error": "Email already registered.", "success": None
        })
    conn.execute(
        "INSERT INTO users (email, password, full_name, father_name, qualifications, experience, role, status) VALUES (?, ?, ?, ?, ?, ?, 'teacher', 'pending')",
        (email.strip().lower(), hash_password(password), full_name.strip(), father_name.strip(), qualifications.strip(), experience.strip()),
    )
    conn.commit()
    conn.close()
    return templates.TemplateResponse("signup.html", {
        "request": request,
        "error": None,
        "success": "Account created! Please wait for admin approval before logging in.",
    })


# =================== ADMIN PANEL (existing) ===================
@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_db()
    pending = conn.execute("SELECT * FROM users WHERE status = 'pending'").fetchall()
    approved_rows = conn.execute(
        "SELECT * FROM users WHERE role = 'teacher' AND status = 'approved' AND is_active = 1 ORDER BY full_name"
    ).fetchall()

    today = date.today().isoformat()
    school_open_today = is_school_open(conn, today)

    approved = []
    for row in approved_rows:
        t = dict(row)
        if not school_open_today:
            t["today_status"] = "holiday"
            t["today_check_in_time"] = None
        else:
            attendance = get_teacher_attendance(conn, row["email"], today)
            if attendance:
                t["today_status"] = "late" if attendance["punctuality"] == "late" else "present"
                t["today_check_in_time"] = attendance["check_in_time"]
            else:
                t["today_status"] = "absent"
                t["today_check_in_time"] = None
        approved.append(t)

    is_open = get_school_open_status(conn, today)
    official_check_in_time = get_official_check_in_time(conn)
    official_check_in_display = format_time_display(official_check_in_time)

    chart_months, chart_values = get_all_teachers_average_trend(conn)
    total_session_open_days = count_total_open_days_marked(conn)

    conn.close()

    return templates.TemplateResponse("admin.html", {
        "request": request,
        "pending": pending,
        "approved": approved,
        "today": today,
        "is_open": is_open,
        "official_check_in_time": official_check_in_time,
        "official_check_in_display": official_check_in_display,
        "user": user,
        "settings_saved": request.query_params.get("settings_saved"),
        "calendar_saved": request.query_params.get("calendar_saved"),
        "chart_months": chart_months,
        "chart_values": chart_values,
        "total_session_open_days": total_session_open_days,
    })


@app.post("/admin/settings/check-in-time")
def update_check_in_time(request: Request, official_check_in_time: str = Form(...)):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_db()
    set_setting(conn, "official_check_in_time", official_check_in_time.strip())
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin?settings_saved=1", status_code=303)


@app.post("/admin/approve/{teacher_id}")
def approve_teacher(teacher_id: int, request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE users SET status = 'approved', is_active = 1 WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)


@app.post("/admin/reject/{teacher_id}")
def reject_teacher(teacher_id: int, request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE users SET status = 'rejected' WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)


@app.post("/admin/remove-teacher/{teacher_id}")
def remove_teacher(teacher_id: int, request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE users SET is_active = 0 WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)


@app.get("/admin/calendar/school-days")
def get_calendar_days(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    conn = get_db()
    events = get_school_calendar_events(conn)
    conn.close()
    return JSONResponse(events)


@app.post("/admin/calendar/school-days")
async def save_calendar_days(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return JSONResponse({"success": False, "error": "unauthorized"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"success": False, "error": "invalid json"}, status_code=400)

    dates = body.get("dates")
    if not isinstance(dates, dict):
        return JSONResponse({"success": False, "error": "invalid dates payload"}, status_code=400)

    conn = get_db()
    bulk_set_school_days(conn, dates)
    conn.close()
    return JSONResponse({"success": True, "count": len(dates)})


@app.get("/admin/teacher/{teacher_id}/reports", response_class=HTMLResponse)
def admin_teacher_reports(teacher_id: int, request: Request, year: int | None = None, month: int | None = None):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_db()
    teacher = conn.execute("SELECT * FROM users WHERE id = ? AND role = 'teacher'", (teacher_id,)).fetchone()
    if not teacher:
        conn.close()
        raise HTTPException(status_code=404, detail="Teacher not found")

    stats = get_progress_stats(conn, teacher["email"], year, month)
    session_totals = get_session_progress_totals(conn, teacher["email"])
    monthly_breakdown = get_monthly_progress_breakdown(conn, teacher["email"])
    conn.close()

    prev_year, prev_month = shift_month(stats["year"], stats["month"], -1)
    next_year, next_month = shift_month(stats["year"], stats["month"], 1)

    return templates.TemplateResponse("admin_teacher_report.html", {
        "request": request,
        "user": user,
        "teacher": teacher,
        "stats": stats,
        "session_totals": session_totals,
        "monthly_breakdown": monthly_breakdown,
        "prev_year": prev_year,
        "prev_month": prev_month,
        "next_year": next_year,
        "next_month": next_month,
    })


# =================== TEACHER DASHBOARD & ATTENDANCE (existing) ===================
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    user = current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user.get("role") == "admin":
        return RedirectResponse(url="/admin", status_code=303)

    today = date.today().isoformat()
    conn = get_db()
    is_open = get_school_open_status(conn, today)
    attendance = get_teacher_attendance(conn, user["email"], today)
    official_check_in_display = format_time_display(get_official_check_in_time(conn))
    conn.close()

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user,
        "today": today,
        "is_open": is_open,
        "attendance": attendance,
        "official_check_in_display": official_check_in_display,
    })


@app.post("/attendance/checkin")
def checkin(request: Request):
    user = current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    today = date.today().isoformat()
    conn = get_db()

    if not is_school_open(conn, today):
        conn.close()
        return RedirectResponse(url="/dashboard?msg=closed", status_code=303)

    existing = get_teacher_attendance(conn, user["email"], today)
    if existing:
        conn.close()
        return RedirectResponse(
            url=f"/dashboard?msg=already_checked&punctuality={existing['punctuality']}",
            status_code=303,
        )

    attendance = record_check_in(conn, user["email"], today)
    conn.close()
    return RedirectResponse(
        url=f"/dashboard?msg=checked_in&punctuality={attendance['punctuality']}",
        status_code=303,
    )


@app.get("/profile", response_class=HTMLResponse)
def profile(request: Request):
    user = current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("profile.html", {"request": request, "user": user})


@app.get("/progress", response_class=HTMLResponse)
def progress(request: Request):
    user = current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user.get("role") == "admin":
        return RedirectResponse(url="/admin", status_code=303)

    conn = get_db()
    stats = get_progress_stats(conn, user["email"])
    session_totals = get_session_progress_totals(conn, user["email"])
    monthly_breakdown = get_monthly_progress_breakdown(conn, user["email"])
    conn.close()

    return templates.TemplateResponse("progress.html", {
        "request": request,
        "user": user,
        "stats": stats,
        "session_totals": session_totals,
        "monthly_breakdown": monthly_breakdown,
    })


# ===================== ADMIN STUDENT MANAGEMENT (NEW) =====================
@app.get("/admin/students", response_class=HTMLResponse)
def admin_students(request: Request):
    user = require_admin(request)
    conn = get_db()
    classes = conn.execute("SELECT * FROM classes ORDER BY name").fetchall()
    templates_info = {}
    for cls in classes:
        tpl = conn.execute(
            "SELECT * FROM class_templates WHERE class_id = ?", (cls["id"],)
        ).fetchone()
        if tpl:
            templates_info[cls["id"]] = {
                "filename": tpl["template_filename"],
                "identity_columns": json.loads(tpl["identity_columns"]),
                "extra_columns": json.loads(tpl["extra_columns"]),
            }
    conn.close()
    return templates.TemplateResponse("admin_students.html", {
        "request": request,
        "user": user,
        "classes": classes,
        "templates_info": templates_info,
    })


@app.post("/admin/students/upload")
async def admin_students_upload(
    request: Request,
    class_id: int = Form(...),
    mode: str = Form("replace"),
    file: UploadFile = File(...),
):
    user = require_admin(request)

    if not file.filename.endswith(('.xlsx', '.xls')):
        return RedirectResponse(url="/admin/students?error=invalid_format", status_code=303)

    suffix = os.path.splitext(file.filename)[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        content = await file.read()
        tmp.write(content)
        tmp.close()

        wb = load_workbook(tmp.name, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            os.unlink(tmp.name)
            return RedirectResponse(url="/admin/students?error=empty_file", status_code=303)

        headers = [str(c).strip() for c in rows[0] if c is not None]
        if len(headers) < 2:
            os.unlink(tmp.name)
            return RedirectResponse(url="/admin/students?error=too_few_columns", status_code=303)

        request.session["tmp_excel_path"] = tmp.name
        request.session["tmp_class_id"] = class_id
        request.session["tmp_mode"] = mode
        request.session["tmp_headers"] = headers

        return templates.TemplateResponse("admin_students_identity.html", {
            "request": request,
            "user": user,
            "class_id": class_id,
            "headers": headers,
        })
    except Exception as e:
        os.unlink(tmp.name)
        return RedirectResponse(url="/admin/students?error=parse_failed", status_code=303)


@app.post("/admin/students/process")
def admin_students_process(
    request: Request,
    identity_col1: str = Form(...),
    identity_col2: str = Form(...),
):
    user = require_admin(request)
    tmp_path = request.session.pop("tmp_excel_path", None)
    class_id = request.session.pop("tmp_class_id", None)
    mode = request.session.pop("tmp_mode", "replace")
    headers = request.session.pop("tmp_headers", [])

    if not tmp_path or not class_id or not os.path.exists(tmp_path):
        return RedirectResponse(url="/admin/students?error=session_expired", status_code=303)

    try:
        wb = load_workbook(tmp_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c).strip() for c in rows[0] if c is not None]

        idx1 = headers.index(identity_col1) if identity_col1 in headers else -1
        idx2 = headers.index(identity_col2) if identity_col2 in headers else -1
        if idx1 == -1 or idx2 == -1:
            os.unlink(tmp_path)
            return RedirectResponse(url="/admin/students?error=invalid_columns", status_code=303)

        extra_cols = [i for i, h in enumerate(headers) if i not in (idx1, idx2)]
        identity_columns = [identity_col1, identity_col2]
        extra_columns = [headers[i] for i in extra_cols]

        conn = get_db()
        if mode == "replace":
            conn.execute("DELETE FROM student_records WHERE class_id = ?", (class_id,))
            conn.execute("DELETE FROM class_templates WHERE class_id = ?", (class_id,))

        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            name = str(row[idx1]).strip() if row[idx1] else ""
            father = str(row[idx2]).strip() if row[idx2] else ""
            if not name or not father:
                continue
            extra_data = {}
            for ci in extra_cols:
                val = row[ci]
                extra_data[headers[ci]] = str(val) if val is not None else ""
            try:
                conn.execute(
                    "INSERT INTO student_records (class_id, name, father_name, extra_data) VALUES (?, ?, ?, ?)",
                    (class_id, name, father, json.dumps(extra_data, ensure_ascii=False)),
                )
            except sqlite3.IntegrityError:
                pass  # duplicate, skip

        template_dir = "uploads/class_templates"
        os.makedirs(template_dir, exist_ok=True)
        perm_path = os.path.join(template_dir, f"class_{class_id}.xlsx")
        shutil.copy(tmp_path, perm_path)

        conn.execute(
            "INSERT OR REPLACE INTO class_templates (class_id, template_filename, identity_columns, extra_columns) VALUES (?, ?, ?, ?)",
            (class_id, perm_path, json.dumps(identity_columns), json.dumps(extra_columns)),
        )
        conn.commit()
        conn.close()

        os.unlink(tmp_path)
        return RedirectResponse(url="/admin/students?msg=uploaded", status_code=303)
    except Exception as e:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        return RedirectResponse(url="/admin/students?error=process_failed", status_code=303)


# ===================== ADMIN MASTER DATA =====================
@app.get("/admin/master-data", response_class=HTMLResponse)
def admin_master_data(request: Request):
    user = require_admin(request)
    conn = get_db()
    subjects = conn.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    test_types = conn.execute("SELECT * FROM test_types ORDER BY name").fetchall()
    conn.close()
    return templates.TemplateResponse("admin_master_data.html", {
        "request": request,
        "user": user,
        "subjects": subjects,
        "test_types": test_types,
    })


@app.post("/admin/master-data/add/subject")
def add_subject(request: Request, name: str = Form(...)):
    user = require_admin(request)
    conn = get_db()
    try:
        conn.execute("INSERT INTO subjects (name) VALUES (?)", (name.strip(),))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@app.post("/admin/master-data/add/test-type")
def add_test_type(request: Request, name: str = Form(...)):
    user = require_admin(request)
    conn = get_db()
    try:
        conn.execute("INSERT INTO test_types (name) VALUES (?)", (name.strip(),))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@app.post("/admin/master-data/delete/subject/{subject_id}")
def delete_subject(subject_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM subjects WHERE id = ?", (subject_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@app.post("/admin/master-data/delete/test-type/{test_type_id}")
def delete_test_type(test_type_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM test_types WHERE id = ?", (test_type_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


# ===================== ADMIN EXAM SESSIONS =====================
@app.get("/admin/exam-sessions", response_class=HTMLResponse)
def admin_exam_sessions(request: Request):
    user = require_admin(request)
    conn = get_db()
    classes = conn.execute("SELECT * FROM classes ORDER BY name").fetchall()
    subjects = conn.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    test_types = conn.execute("SELECT * FROM test_types ORDER BY name").fetchall()
    teachers = conn.execute(
        "SELECT email, full_name FROM users WHERE role = 'teacher' AND status = 'approved' AND is_active = 1 ORDER BY full_name"
    ).fetchall()
    sessions = conn.execute("""
        SELECT es.*, c.name as class_name, tt.name as test_type_name,
        (SELECT COUNT(*) FROM exam_session_subjects WHERE session_id = es.id) as subject_count
        FROM exam_sessions es
        JOIN classes c ON c.id = es.class_id
        JOIN test_types tt ON tt.id = es.test_type_id
        ORDER BY es.created_at DESC
    """).fetchall()
    conn.close()
    return templates.TemplateResponse("admin_exam_sessions.html", {
        "request": request,
        "user": user,
        "classes": classes,
        "subjects": subjects,
        "test_types": test_types,
        "teachers": teachers,
        "sessions": sessions,
    })


@app.post("/admin/exam-sessions/create")
async def admin_create_exam_session(request: Request):
    user = require_admin(request)
    form = await request.form()
    class_id = int(form["class_id"])
    test_type_id = int(form["test_type_id"])
    test_number = form["test_number"].strip()
    conduct_date = form["conduct_date"].strip()
    syllabus = form.get("syllabus", "").strip()
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO exam_sessions (class_id, test_type_id, test_number, conduct_date, syllabus)
        VALUES (?, ?, ?, ?, ?)
    """, (class_id, test_type_id, test_number, conduct_date, syllabus))
    session_id = cur.lastrowid

    idx = 0
    while True:
        subj_key = f"subject_id_{idx}"
        if subj_key not in form:
            break
        subject_id = int(form[subj_key])
        teacher_email = form[f"teacher_email_{idx}"].strip()
        total_marks = float(form[f"total_marks_{idx}"])
        passing_marks = float(form[f"passing_marks_{idx}"])
        cur.execute("""
            INSERT INTO exam_session_subjects (session_id, subject_id, teacher_email, total_marks, passing_marks)
            VALUES (?, ?, ?, ?, ?)
        """, (session_id, subject_id, teacher_email, total_marks, passing_marks))
        idx += 1

    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-sessions?msg=created", status_code=303)


# ===================== TEACHER EXAM MARKS ENTRY (NEW) =====================
@app.get("/test-marks", response_class=HTMLResponse)
def teacher_exam_marks_home(request: Request, session_id: int | None = None):
    user = require_login(request)
    if user.get("role") == "admin":
        return RedirectResponse(url="/admin/exam-status", status_code=303)

    conn = get_db()
    sessions = conn.execute("""
        SELECT DISTINCT es.id, c.name as class_name, tt.name as test_type_name,
               es.test_number, es.conduct_date
        FROM exam_sessions es
        JOIN exam_session_subjects ess ON ess.session_id = es.id
        JOIN classes c ON c.id = es.class_id
        JOIN test_types tt ON tt.id = es.test_type_id
        WHERE ess.teacher_email = ?
        ORDER BY es.conduct_date DESC
    """, (user["email"],)).fetchall()

    selected_session_id = session_id or (sessions[0]["id"] if sessions else None)
    session_subjects = []
    if selected_session_id:
        session_subjects = conn.execute("""
            SELECT ess.*, s.name as subject_name
            FROM exam_session_subjects ess
            JOIN subjects s ON s.id = ess.subject_id
            WHERE ess.session_id = ? AND ess.teacher_email = ?
        """, (selected_session_id, user["email"])).fetchall()
    conn.close()
    return templates.TemplateResponse("teacher_exam_marks.html", {
        "request": request,
        "user": user,
        "sessions": sessions,
        "selected_session_id": selected_session_id,
        "session_subjects": session_subjects,
    })


@app.get("/test-marks/entry", response_class=HTMLResponse)
def marks_entry_grid(request: Request, session_subject_id: int):
    user = require_login(request)
    conn = get_db()
    ss = conn.execute("""
        SELECT ess.*, s.name as subject_name, es.class_id, es.test_type_id, es.test_number,
               tt.name as test_type_name, c.name as class_name
        FROM exam_session_subjects ess
        JOIN exam_sessions es ON es.id = ess.session_id
        JOIN subjects s ON s.id = ess.subject_id
        JOIN test_types tt ON tt.id = es.test_type_id
        JOIN classes c ON c.id = es.class_id
        WHERE ess.id = ?
    """, (session_subject_id,)).fetchone()
    if not ss or ss["teacher_email"] != user["email"]:
        conn.close()
        return RedirectResponse(url="/test-marks?error=unauthorized", status_code=303)

    tpl = conn.execute("SELECT * FROM class_templates WHERE class_id = ?", (ss["class_id"],)).fetchone()
    extra_columns = json.loads(tpl["extra_columns"]) if tpl else []
    identity_columns = json.loads(tpl["identity_columns"]) if tpl else ["Name", "Father Name"]

    students = conn.execute("""
        SELECT sr.*, em.marks_obtained as mark
        FROM student_records sr
        LEFT JOIN exam_marks em ON em.student_id = sr.id
            AND em.session_subject_id = ?
        WHERE sr.class_id = ?
        ORDER BY sr.name
    """, (session_subject_id, ss["class_id"])).fetchall()

    conn.close()
    return templates.TemplateResponse("teacher_exam_marks_grid.html", {
        "request": request,
        "user": user,
        "session_subject_id": session_subject_id,
        "session_id": ss["session_id"],
        "subject_name": ss["subject_name"],
        "class_name": ss["class_name"],
        "test_info": f"{ss['test_type_name']} {ss['test_number']}",
        "total_marks": ss["total_marks"],
        "passing_marks": ss["passing_marks"],
        "students": students,
        "extra_columns": extra_columns,
    })


@app.post("/test-marks/save-new")
async def save_marks_new(request: Request):
    user = require_login(request)
    form = await request.form()
    session_subject_id = int(form["session_subject_id"])
    session_id = int(form["session_id"])

    conn = get_db()
    ss = conn.execute("SELECT * FROM exam_session_subjects WHERE id = ? AND teacher_email = ?",
                      (session_subject_id, user["email"])).fetchone()
    if not ss:
        conn.close()
        return RedirectResponse(url="/test-marks?error=unauthorized", status_code=303)

    conn.execute("DELETE FROM exam_marks WHERE session_subject_id = ?", (session_subject_id,))

    for key in form.keys():
        if key.startswith("mark_"):
            student_id = int(key.split("_")[1])
            raw = form[key]
            if raw and str(raw).strip():
                marks_obtained = float(raw)
                conn.execute("INSERT INTO exam_marks (session_subject_id, student_id, marks_obtained) VALUES (?, ?, ?)",
                             (session_subject_id, student_id, marks_obtained))

    conn.execute("UPDATE exam_session_subjects SET submitted = 1 WHERE id = ?", (session_subject_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/test-marks?msg=saved&session_id={session_id}", status_code=303)


# ===================== ADMIN EXAM STATUS & CONFIRMATION =====================
@app.get("/admin/exam-status", response_class=HTMLResponse)
def admin_exam_status(request: Request):
    user = require_admin(request)
    conn = get_db()
    session_rows = conn.execute("""
        SELECT es.*, c.name as class_name, tt.name as test_type_name
        FROM exam_sessions es
        JOIN classes c ON c.id = es.class_id
        JOIN test_types tt ON tt.id = es.test_type_id
        ORDER BY es.conduct_date DESC
    """).fetchall()

    sessions = []
    for sess in session_rows:
        subjects = conn.execute("""
            SELECT ess.*, s.name as subject_name, u.full_name as teacher_name
            FROM exam_session_subjects ess
            JOIN subjects s ON s.id = ess.subject_id
            JOIN users u ON u.email = ess.teacher_email
            WHERE ess.session_id = ?
        """, (sess["id"],)).fetchall()
        sess_dict = dict(sess)
        sess_dict["subjects"] = [dict(s) for s in subjects]
        sessions.append(sess_dict)
    conn.close()
    return templates.TemplateResponse("admin_exam_status.html", {
        "request": request,
        "user": user,
        "sessions": sessions,
    })


@app.post("/admin/exam-status/confirm/{session_id}")
def confirm_result(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("UPDATE exam_sessions SET status = 'confirmed' WHERE id = ?", (session_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-status?msg=confirmed", status_code=303)


@app.post("/admin/exam-status/reopen/{session_id}")
def reopen_session(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("UPDATE exam_sessions SET status = 'open' WHERE id = ?", (session_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-status?msg=reopened", status_code=303)


# ===================== ADMIN EXPORT SESSION EXCEL =====================
@app.get("/admin/exam-session/{session_id}/export")
def export_session_excel(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    session = conn.execute("""
        SELECT es.*, c.name as class_name
        FROM exam_sessions es JOIN classes c ON c.id = es.class_id
        WHERE es.id = ?
    """, (session_id,)).fetchone()
    if not session:
        conn.close()
        raise HTTPException(status_code=404)

    subjects = conn.execute("""
        SELECT ess.*, s.name as subject_name
        FROM exam_session_subjects ess JOIN subjects s ON s.id = ess.subject_id
        WHERE ess.session_id = ?
    """, (session_id,)).fetchall()

    tpl = conn.execute("SELECT * FROM class_templates WHERE class_id = ?", (session["class_id"],)).fetchone()
    identity_cols = json.loads(tpl["identity_columns"]) if tpl else ["Name", "Father Name"]

    students = conn.execute("""
        SELECT * FROM student_records WHERE class_id = ? ORDER BY name
    """, (session["class_id"],)).fetchall()

    if tpl and os.path.exists(tpl["template_filename"]):
        wb = load_workbook(tpl["template_filename"])
        ws = wb.active
        max_col = ws.max_column
        col_offset = max_col + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        for i, col in enumerate(identity_cols, start=1):
            ws.cell(row=1, column=i, value=col).font = Font(bold=True)
        col_offset = len(identity_cols) + 1

    for j, subj in enumerate(subjects):
        header = f"{subj['subject_name']} (Out of {subj['total_marks']})"
        ws.cell(row=1, column=col_offset + j, value=header).font = Font(bold=True)

    for i, student in enumerate(students, start=2):
        if not tpl or not os.path.exists(tpl["template_filename"]):
            ws.cell(row=i, column=1, value=student["name"])
            ws.cell(row=i, column=2, value=student["father_name"])
        for j, subj in enumerate(subjects):
            mark_row = conn.execute("""
                SELECT marks_obtained FROM exam_marks
                WHERE session_subject_id = ? AND student_id = ?
            """, (subj["id"], student["id"])).fetchone()
            mark_val = mark_row["marks_obtained"] if mark_row else ""
            ws.cell(row=i, column=col_offset + j, value=mark_val)

    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"result_{session['class_name']}_{session['test_number']}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )