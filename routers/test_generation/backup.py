import io
from datetime import date

from fastapi import APIRouter, Request, File, UploadFile, Form
from fastapi.responses import StreamingResponse, RedirectResponse

from database import get_db
from main import templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from services.sanitizer import clean_html
from .utils import require_test_gen_access

router = APIRouter(prefix="/test-generation")

def export_questions_to_excel(block_ids=None):
    """Create Excel file with question bank data. If block_ids is None, export all."""
    conn = get_db()
    if block_ids:
        placeholders = ','.join('?' for _ in block_ids)
        query = f"""
            SELECT q.*, sb.block_name, c.name as class_name, s.name as subject_name
            FROM question_bank q
            JOIN syllabus_blocks sb ON sb.id = q.syllabus_block_id
            JOIN classes c ON c.id = sb.class_id
            JOIN subjects s ON s.id = sb.subject_id
            WHERE q.syllabus_block_id IN ({placeholders})
            ORDER BY sb.block_name, q.question_type, q.id
        """
        rows = conn.execute(query, block_ids).fetchall()
    else:
        rows = conn.execute("""
            SELECT q.*, sb.block_name, c.name as class_name, s.name as subject_name
            FROM question_bank q
            JOIN syllabus_blocks sb ON sb.id = q.syllabus_block_id
            JOIN classes c ON c.id = sb.class_id
            JOIN subjects s ON s.id = sb.subject_id
            ORDER BY sb.block_name, q.question_type, q.id
        """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "QuestionBank"

    # Headers include parent_comprehension_id to preserve links
    headers = [
        "Block Name", "Class", "Subject", "Type", "Question",
        "Option A", "Option B", "Option C", "Option D", "Correct",
        "Answer", "Passage", "ParentComprehensionId"
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    row_num = 2
    for r in rows:
        ws.cell(row=row_num, column=1, value=r["block_name"])
        ws.cell(row=row_num, column=2, value=r["class_name"])
        ws.cell(row=row_num, column=3, value=r["subject_name"])
        ws.cell(row=row_num, column=4, value=r["question_type"])
        ws.cell(row=row_num, column=5, value=r["question_text"])
        ws.cell(row=row_num, column=6, value=r["option_a"] or "")
        ws.cell(row=row_num, column=7, value=r["option_b"] or "")
        ws.cell(row=row_num, column=8, value=r["option_c"] or "")
        ws.cell(row=row_num, column=9, value=r["option_d"] or "")
        ws.cell(row=row_num, column=10, value=r["correct_answer"] or "")
        ws.cell(row=row_num, column=11, value=r["answer_text"] or "")
        ws.cell(row=row_num, column=12, value=r["comprehension_passage"] or "")
        # ParentComprehensionId (empty if none)
        parent_id = r["parent_comprehension_id"] if r["parent_comprehension_id"] else ""
        ws.cell(row=row_num, column=13, value=parent_id)
        row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_questions_from_excel(file, mode="append", block_id=None):
    """
    Read Excel file and import questions.
    If block_id is given, restrict to that block (replace/append within block).
    If block_id is None, process all rows (global replace/append).
    """
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    headers = [str(c).strip() if c else "" for c in rows[0]]
    required = ["Block Name", "Class", "Subject", "Type", "Question"]
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing column: {col}")

    idx = {col: headers.index(col) for col in required}
    # Optional columns
    opt_idx = {}
    for col in ["Option A", "Option B", "Option C", "Option D", "Correct", "Answer", "Passage", "ParentComprehensionId"]:
        if col in headers:
            opt_idx[col] = headers.index(col)

    conn = get_db()
    cur = conn.cursor()

    # If block_id is given and mode replace, delete existing questions for that block
    if block_id and mode == "replace":
        cur.execute("DELETE FROM question_bank WHERE syllabus_block_id = ?", (block_id,))
        conn.commit()

    # For global replace, we need to delete all questions? We'll handle later.
    # For now, implement append logic per row.

    inserted_count = 0
    # We need to keep track of comprehension IDs for linking MCQs later.
    # Because passage rows are before MCQs? We'll process rows sequentially and map parent IDs.
    # ParentComprehensionId refers to the original ID from export, not new IDs.
    # So we need a mapping from old ID to new ID. We'll store old_id -> new_id for comprehension rows.

    old_to_new = {}

    for row in rows[1:]:
        if not any(row):
            continue
        block_name = str(row[idx["Block Name"]]).strip() if row[idx["Block Name"]] else ""
        class_name = str(row[idx["Class"]]).strip() if row[idx["Class"]] else ""
        subject_name = str(row[idx["Subject"]]).strip() if row[idx["Subject"]] else ""
        q_type = str(row[idx["Type"]]).strip().lower() if row[idx["Type"]] else ""
        question_text = clean_html(str(row[idx["Question"]])) if row[idx["Question"]] else ""

        if not question_text:
            continue  # skip empty

        # Find or create syllabus block
        block_row = cur.execute("""
            SELECT sb.id FROM syllabus_blocks sb
            JOIN classes c ON c.id = sb.class_id
            JOIN subjects s ON s.id = sb.subject_id
            WHERE sb.block_name = ? AND c.name = ? AND s.name = ?
        """, (block_name, class_name, subject_name)).fetchone()
        if block_row:
            new_block_id = block_row["id"]
        else:
            # Create class/subject if not exist? We assume they exist in system.
            # For simplicity, use existing class/subject names to find IDs.
            class_row = cur.execute("SELECT id FROM classes WHERE name = ?", (class_name,)).fetchone()
            subject_row = cur.execute("SELECT id FROM subjects WHERE name = ?", (subject_name,)).fetchone()
            if not class_row or not subject_row:
                continue  # skip invalid
            cur.execute("INSERT INTO syllabus_blocks (class_id, subject_id, block_name) VALUES (?, ?, ?)",
                        (class_row["id"], subject_row["id"], block_name))
            new_block_id = cur.lastrowid
            block_row = {"id": new_block_id}

        # If block_id specified and this row belongs to different block, skip
        if block_id and new_block_id != block_id:
            continue

        # Extract optional fields
        opt_a = clean_html(str(row[opt_idx["Option A"]])) if "Option A" in opt_idx and row[opt_idx["Option A"]] else None
        opt_b = clean_html(str(row[opt_idx["Option B"]])) if "Option B" in opt_idx and row[opt_idx["Option B"]] else None
        opt_c = clean_html(str(row[opt_idx["Option C"]])) if "Option C" in opt_idx and row[opt_idx["Option C"]] else None
        opt_d = clean_html(str(row[opt_idx["Option D"]])) if "Option D" in opt_idx and row[opt_idx["Option D"]] else None
        correct = str(row[opt_idx["Correct"]]).strip() if "Correct" in opt_idx and row[opt_idx["Correct"]] else None
        answer = clean_html(str(row[opt_idx["Answer"]])) if "Answer" in opt_idx and row[opt_idx["Answer"]] else None
        passage = clean_html(str(row[opt_idx["Passage"]])) if "Passage" in opt_idx and row[opt_idx["Passage"]] else None
        parent_old = row[opt_idx["ParentComprehensionId"]] if "ParentComprehensionId" in opt_idx and row[opt_idx["ParentComprehensionId"]] else None
        if parent_old:
            parent_old = int(parent_old)
        else:
            parent_old = None

        # For comprehension passage rows (q_type == 'comprehension'), insert and get new ID
        if q_type == "comprehension":
            cur.execute("""
                INSERT INTO question_bank (syllabus_block_id, question_type, comprehension_passage)
                VALUES (?, 'comprehension', ?)
            """, (new_block_id, passage or ""))
            new_id = cur.lastrowid
            if parent_old is None:
                # This is a new comprehension row; map old ID? we don't have old ID unless we add it.
                # We'll skip for now; but we need to handle MCQ linking later.
                pass
            continue

        # For MCQ or short/long
        if q_type in ("mcq", "short", "long"):
            # Check duplicate by exact question_text within this block
            existing = cur.execute("""
                SELECT id FROM question_bank
                WHERE syllabus_block_id = ? AND question_text = ? AND question_type = ?
            """, (new_block_id, question_text, q_type)).fetchone()
            if existing and mode == "append":
                continue  # skip duplicate
            if existing and mode == "replace":
                # Update existing
                cur.execute("""
                    UPDATE question_bank
                    SET option_a=?, option_b=?, option_c=?, option_d=?, correct_answer=?, answer_text=?, parent_comprehension_id=?
                    WHERE id=?
                """, (opt_a, opt_b, opt_c, opt_d, correct, answer, parent_old if parent_old else None, existing["id"]))
                continue
            else:
                cur.execute("""
                    INSERT INTO question_bank (syllabus_block_id, question_type, question_text,
                                              option_a, option_b, option_c, option_d, correct_answer,
                                              answer_text, parent_comprehension_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (new_block_id, q_type, question_text, opt_a, opt_b, opt_c, opt_d, correct, answer, parent_old if parent_old else None))
                inserted_count += 1

    conn.commit()
    conn.close()
    return inserted_count


@router.get("/download-block/{block_id}")
def download_block(block_id: int, request: Request):
    user = require_test_gen_access(request)
    output = export_questions_to_excel([block_id])
    filename = f"question_bank_block_{block_id}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/download-all")
def download_all(request: Request):
    user = require_test_gen_access(request)
    output = export_questions_to_excel(None)
    filename = f"question_bank_all_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/upload-block/{block_id}")
async def upload_block(
    block_id: int,
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form("append"),
):
    user = require_test_gen_access(request)
    contents = await file.read()
    try:
        import_questions_from_excel(io.BytesIO(contents), mode=mode, block_id=block_id)
    except Exception as e:
        return RedirectResponse(url=f"/test-generation/question-bank?error=upload_failed", status_code=303)
    return RedirectResponse(url=f"/test-generation/question-bank?msg=uploaded", status_code=303)


@router.post("/upload-all")
async def upload_all(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form("append"),
):
    user = require_test_gen_access(request)
    contents = await file.read()
    try:
        import_questions_from_excel(io.BytesIO(contents), mode=mode, block_id=None)
    except Exception as e:
        return RedirectResponse(url=f"/test-generation/question-bank?error=upload_failed", status_code=303)
    return RedirectResponse(url=f"/test-generation/question-bank?msg=uploaded", status_code=303)