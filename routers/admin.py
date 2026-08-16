import json, os, shutil, tempfile
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from database import (
    get_db,
    init_db,
    get_official_check_in_time,
    format_time_display,
    get_school_open_status,
    is_school_open,
    get_teacher_attendance,
    get_progress_stats,
    get_session_progress_totals,
    get_monthly_progress_breakdown,
    get_all_teachers_average_trend,
    bulk_set_school_days,
    get_school_calendar_events,
    count_total_open_days_marked,
    set_setting,
    get_setting,
)

router = APIRouter()

from main import current_user, require_admin, require_login, shift_month, templates


# =================== ADMIN PANEL ===================
@router.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_db()
    pending = conn.execute("SELECT * FROM users WHERE status = 'pending'").fetchall()
    approved_rows = conn.execute(
        "SELECT * FROM users WHERE role = 'teacher' AND status = 'approved' AND is_active = 1 ORDER BY full_name"
    ).fetchall()

    today = date.today().isoformat()
    school_open_today = is_school_open(conn, today)

    approved = []
    for row in approved_rows:
        t = dict(row)
        if not school_open_today:
            t["today_status"] = "holiday"
            t["today_check_in_time"] = None
        else:
            attendance = get_teacher_attendance(conn, row["email"], today)
            if attendance:
                t["today_status"] = "late" if attendance["punctuality"] == "late" else "present"
                t["today_check_in_time"] = attendance["check_in_time"]
            else:
                t["today_status"] = "absent"
                t["today_check_in_time"] = None
        approved.append(t)

    is_open = get_school_open_status(conn, today)
    official_check_in_time = get_official_check_in_time(conn)
    official_check_in_display = format_time_display(official_check_in_time)

    chart_months, chart_values = get_all_teachers_average_trend(conn)
    total_session_open_days = count_total_open_days_marked(conn)

    conn.close()

    return templates.TemplateResponse("admin.html", {
        "request": request,
        "pending": pending,
        "approved": approved,
        "today": today,
        "is_open": is_open,
        "official_check_in_time": official_check_in_time,
        "official_check_in_display": official_check_in_display,
        "user": user,
        "settings_saved": request.query_params.get("settings_saved"),
        "calendar_saved": request.query_params.get("calendar_saved"),
        "chart_months": chart_months,
        "chart_values": chart_values,
        "total_session_open_days": total_session_open_days,
    })


@router.post("/admin/settings/check-in-time")
def update_check_in_time(request: Request, official_check_in_time: str = Form(...)):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    set_setting(conn, "official_check_in_time", official_check_in_time.strip())
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin?settings_saved=1", status_code=303)


@router.post("/admin/approve/{teacher_id}")
def approve_teacher(teacher_id: int, request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE users SET status = 'approved', is_active = 1 WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)


@router.post("/admin/reject/{teacher_id}")
def reject_teacher(teacher_id: int, request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE users SET status = 'rejected' WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)


@router.post("/admin/remove-teacher/{teacher_id}")
def remove_teacher(teacher_id: int, request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE users SET is_active = 0 WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)


@router.get("/admin/calendar/school-days")
def get_calendar_days(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    conn = get_db()
    events = get_school_calendar_events(conn)
    conn.close()
    return JSONResponse(events)


@router.post("/admin/calendar/school-days")
async def save_calendar_days(request: Request):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return JSONResponse({"success": False, "error": "unauthorized"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"success": False, "error": "invalid json"}, status_code=400)

    dates = body.get("dates")
    if not isinstance(dates, dict):
        return JSONResponse({"success": False, "error": "invalid dates payload"}, status_code=400)

    conn = get_db()
    bulk_set_school_days(conn, dates)
    conn.close()
    return JSONResponse({"success": True, "count": len(dates)})


@router.get("/admin/teacher/{teacher_id}/reports", response_class=HTMLResponse)
def admin_teacher_reports(teacher_id: int, request: Request, year: int | None = None, month: int | None = None):
    user = current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_db()
    teacher = conn.execute("SELECT * FROM users WHERE id = ? AND role = 'teacher'", (teacher_id,)).fetchone()
    if not teacher:
        conn.close()
        raise HTTPException(status_code=404, detail="Teacher not found")

    stats = get_progress_stats(conn, teacher["email"], year, month)
    session_totals = get_session_progress_totals(conn, teacher["email"])
    monthly_breakdown = get_monthly_progress_breakdown(conn, teacher["email"])
    conn.close()

    prev_year, prev_month = shift_month(stats["year"], stats["month"], -1)
    next_year, next_month = shift_month(stats["year"], stats["month"], 1)

    return templates.TemplateResponse("admin_teacher_report.html", {
        "request": request,
        "user": user,
        "teacher": teacher,
        "stats": stats,
        "session_totals": session_totals,
        "monthly_breakdown": monthly_breakdown,
        "prev_year": prev_year,
        "prev_month": prev_month,
        "next_year": next_year,
        "next_month": next_month,
    })


# ===================== ADMIN STUDENT MANAGEMENT =====================
@router.get("/admin/students", response_class=HTMLResponse)
def admin_students(request: Request, class_id: int | None = None):
    user = require_admin(request)
    conn = get_db()
    classes = conn.execute("SELECT * FROM classes ORDER BY name").fetchall()
    templates_info = {}
    for cls in classes:
        tpl = conn.execute(
            "SELECT * FROM class_templates WHERE class_id = ?", (cls["id"],)
        ).fetchone()
        if tpl:
            templates_info[cls["id"]] = {
                "filename": tpl["template_filename"],
                "identity_columns": json.loads(tpl["identity_columns"]),
                "extra_columns": json.loads(tpl["extra_columns"]),
            }

    selected_class_id = class_id or (classes[0]["id"] if classes else None)
    students = []
    if selected_class_id:
        students_raw = conn.execute("""
            SELECT * FROM student_records WHERE class_id = ? ORDER BY name
        """, (selected_class_id,)).fetchall()
        students = [dict(s) for s in students_raw]

    conn.close()
    return templates.TemplateResponse("admin_students.html", {
        "request": request,
        "user": user,
        "classes": classes,
        "templates_info": templates_info,
        "selected_class_id": selected_class_id,
        "students": students,
    })


@router.post("/admin/students/upload")
async def admin_students_upload(
    request: Request,
    class_id: int = Form(...),
    mode: str = Form("replace"),
    file: UploadFile = File(...),
):
    user = require_admin(request)

    if not file.filename.endswith(('.xlsx', '.xls')):
        return RedirectResponse(url="/admin/students?error=invalid_format", status_code=303)

    suffix = os.path.splitext(file.filename)[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        content = await file.read()
        tmp.write(content)
        tmp.close()

        wb = load_workbook(tmp.name, read_only=False)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            os.unlink(tmp.name)
            return RedirectResponse(url="/admin/students?error=empty_file", status_code=303)

        clean_rows = []
        for row in rows:
            clean_row = []
            for cell in row:
                clean_row.append(str(cell).strip() if cell is not None else "")
            clean_rows.append(clean_row)

        header_idx = -1
        name_col = -1
        father_col = -1
        phone_col = -1
        for i, row in enumerate(clean_rows):
            lowered = [c.lower() for c in row]
            if "student name" in lowered and "father name" in lowered:
                header_idx = i
                name_col = lowered.index("student name")
                father_col = lowered.index("father name")
                # Detect phone column (case-insensitive)
                for j, h in enumerate(lowered):
                    if any(key in h for key in ["phone", "mobile", "contact"]):
                        phone_col = j
                        break
                break

        if header_idx == -1:
            os.unlink(tmp.name)
            return RedirectResponse(url="/admin/students?error=headers_not_found", status_code=303)

        extra_cols = []
        for idx, h in enumerate(clean_rows[header_idx]):
            if idx == name_col or idx == father_col or idx == phone_col:
                continue
            if h:
                extra_cols.append((idx, h))

        conn = get_db()
        if mode == "replace":
            conn.execute("DELETE FROM student_records WHERE class_id = ?", (class_id,))
            conn.execute("DELETE FROM class_templates WHERE class_id = ?", (class_id,))

        for row in clean_rows[header_idx+1:]:
            name = row[name_col].strip() if name_col < len(row) else ""
            father = row[father_col].strip() if father_col < len(row) else ""
            if not name or not father:
                continue
            phone = ""
            if phone_col != -1 and phone_col < len(row):
                phone = str(row[phone_col]).strip()
                # Normalize phone to +92XXXXXXXXXX
                phone = phone.replace(" ", "").replace("-", "")
                if phone and not phone.startswith("+92"):
                    if phone.startswith("0"):
                        phone = "+92" + phone[1:]
                    else:
                        phone = "+92" + phone
            extra_data = {}
            for idx, h in extra_cols:
                val = row[idx] if idx < len(row) else ""
                extra_data[h] = val
            try:
                conn.execute(
                    "INSERT INTO student_records (class_id, name, father_name, phone, extra_data) VALUES (?, ?, ?, ?, ?)",
                    (class_id, name, father, phone, json.dumps(extra_data, ensure_ascii=False)),
                )
            except sqlite3.IntegrityError:
                pass

        template_dir = "uploads/class_templates"
        os.makedirs(template_dir, exist_ok=True)
        perm_path = os.path.join(template_dir, f"class_{class_id}.xlsx")
        shutil.copy(tmp.name, perm_path)

        identity_columns = [clean_rows[header_idx][name_col], clean_rows[header_idx][father_col]]
        extra_column_names = [h for _, h in extra_cols]

        conn.execute(
            "INSERT OR REPLACE INTO class_templates (class_id, template_filename, identity_columns, extra_columns) VALUES (?, ?, ?, ?)",
            (class_id, perm_path, json.dumps(identity_columns), json.dumps(extra_column_names)),
        )
        conn.commit()
        conn.close()

        os.unlink(tmp.name)
        return RedirectResponse(url="/admin/students?msg=uploaded", status_code=303)
    except Exception as e:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)
        return RedirectResponse(url="/admin/students?error=parse_failed", status_code=303)


@router.get("/admin/students/edit", response_class=HTMLResponse)
def admin_students_edit(request: Request, class_id: int):
    user = require_admin(request)
    conn = get_db()
    cls = conn.execute("SELECT * FROM classes WHERE id = ?", (class_id,)).fetchone()
    if not cls:
        conn.close()
        return RedirectResponse(url="/admin/students?error=class_not_found", status_code=303)

    tpl = conn.execute("SELECT * FROM class_templates WHERE class_id = ?", (class_id,)).fetchone()
    extra_columns = json.loads(tpl["extra_columns"]) if tpl else []
    identity_columns = json.loads(tpl["identity_columns"]) if tpl else ["Name", "Father Name"]

    students_raw = conn.execute("""
        SELECT * FROM student_records WHERE class_id = ? ORDER BY name
    """, (class_id,)).fetchall()
    conn.close()

    students = []
    for s in students_raw:
        s_dict = dict(s)
        try:
            s_dict['extra_data'] = json.loads(s_dict['extra_data']) if s_dict['extra_data'] else {}
        except (json.JSONDecodeError, TypeError):
            s_dict['extra_data'] = {}
        students.append(s_dict)

    return templates.TemplateResponse("admin_students_edit.html", {
        "request": request,
        "user": user,
        "class": dict(cls),
        "class_id": class_id,
        "students": students,
        "extra_columns": extra_columns,
        "identity_columns": identity_columns,
    })


@router.post("/admin/students/update")
async def admin_students_update(request: Request):
    user = require_admin(request)
    form = await request.form()
    student_id = int(form.get("student_id"))
    name = form.get("name", "").strip()
    father = form.get("father_name", "").strip()
    class_id = int(form.get("class_id", 0))
    phone = form.get("phone", "").strip()
    phone = phone.replace(" ", "").replace("-", "")
    if phone and not phone.startswith("+92"):
        if phone.startswith("0"):
            phone = "+92" + phone[1:]
        else:
            phone = "+92" + phone

    extra_data = {}
    for key in form.keys():
        if key.startswith("extra_data[") and key.endswith("]"):
            col_name = key[len("extra_data["):-1]
            extra_data[col_name] = form[key].strip()

    conn = get_db()
    conn.execute("""
        UPDATE student_records SET name = ?, father_name = ?, phone = ?, extra_data = ?
        WHERE id = ?
    """, (name, father, phone, json.dumps(extra_data, ensure_ascii=False), student_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/admin/students/edit?class_id={class_id}&msg=updated", status_code=303)


@router.post("/admin/students/add")
async def admin_students_add(request: Request):
    user = require_admin(request)
    form = await request.form()
    class_id = int(form.get("class_id"))
    name = form.get("name", "").strip()
    father = form.get("father_name", "").strip()
    phone = form.get("phone", "").strip()
    phone = phone.replace(" ", "").replace("-", "")
    if phone and not phone.startswith("+92"):
        if phone.startswith("0"):
            phone = "+92" + phone[1:]
        else:
            phone = "+92" + phone

    extra_data = {}
    for key in form.keys():
        if key.startswith("extra_data[") and key.endswith("]"):
            col_name = key[len("extra_data["):-1]
            extra_data[col_name] = form[key].strip()

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO student_records (class_id, name, father_name, phone, extra_data)
            VALUES (?, ?, ?, ?, ?)
        """, (class_id, name, father, phone, json.dumps(extra_data, ensure_ascii=False)))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url=f"/admin/students/edit?class_id={class_id}&msg=added", status_code=303)


@router.post("/admin/students/delete/{student_id}")
def admin_students_delete(student_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    row = conn.execute("SELECT class_id FROM student_records WHERE id = ?", (student_id,)).fetchone()
    if row:
        class_id = row["class_id"]
        conn.execute("DELETE FROM student_records WHERE id = ?", (student_id,))
        conn.commit()
    conn.close()
    return RedirectResponse(url=f"/admin/students/edit?class_id={class_id}&msg=deleted", status_code=303)


# ===================== ADMIN MASTER DATA =====================
@router.get("/admin/master-data", response_class=HTMLResponse)
def admin_master_data(request: Request):
    user = require_admin(request)
    conn = get_db()
    classes = conn.execute("SELECT * FROM classes ORDER BY name").fetchall()
    subjects = conn.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    test_types = conn.execute("SELECT * FROM test_types ORDER BY name").fetchall()
    conn.close()
    return templates.TemplateResponse("admin_master_data.html", {
        "request": request,
        "user": user,
        "classes": classes,
        "subjects": subjects,
        "test_types": test_types,
    })


@router.post("/admin/master-data/add/subject")
def add_subject(request: Request, name: str = Form(...)):
    user = require_admin(request)
    conn = get_db()
    try:
        conn.execute("INSERT INTO subjects (name, default_total_marks, default_passing_marks) VALUES (?, 25, 10)", (name.strip(),))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@router.post("/admin/master-data/update/subject/{subject_id}")
async def update_subject_defaults(subject_id: int, request: Request):
    user = require_admin(request)
    form = await request.form()
    total = float(form.get("default_total_marks", 25))
    passing = float(form.get("default_passing_marks", 10))
    conn = get_db()
    conn.execute("UPDATE subjects SET default_total_marks = ?, default_passing_marks = ? WHERE id = ?",
                 (total, passing, subject_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@router.post("/admin/master-data/add/test-type")
def add_test_type(request: Request, name: str = Form(...)):
    user = require_admin(request)
    conn = get_db()
    try:
        conn.execute("INSERT INTO test_types (name) VALUES (?)", (name.strip(),))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@router.post("/admin/master-data/add/class")
def add_class(request: Request, name: str = Form(...)):
    user = require_admin(request)
    conn = get_db()
    try:
        conn.execute("INSERT INTO classes (name) VALUES (?)", (name.strip(),))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@router.post("/admin/master-data/delete/subject/{subject_id}")
def delete_subject(subject_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM subjects WHERE id = ?", (subject_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@router.post("/admin/master-data/delete/test-type/{test_type_id}")
def delete_test_type(test_type_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM test_types WHERE id = ?", (test_type_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


@router.post("/admin/master-data/delete/class/{class_id}")
def delete_class(class_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM classes WHERE id = ?", (class_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/master-data", status_code=303)


# ===================== ADMIN EXAM SESSIONS =====================
@router.get("/admin/exam-sessions", response_class=HTMLResponse)
def admin_exam_sessions(request: Request):
    user = require_admin(request)
    conn = get_db()
    classes = conn.execute("SELECT * FROM classes ORDER BY name").fetchall()
    subjects = conn.execute("SELECT id, name, default_total_marks, default_passing_marks FROM subjects ORDER BY name").fetchall()
    test_types = conn.execute("SELECT * FROM test_types ORDER BY name").fetchall()
    sessions = conn.execute("""
        SELECT es.*, c.name as class_name, tt.name as test_type_name,
        (SELECT COUNT(*) FROM exam_session_subjects WHERE session_id = es.id) as subject_count
        FROM exam_sessions es
        JOIN classes c ON c.id = es.class_id
        JOIN test_types tt ON tt.id = es.test_type_id
        ORDER BY es.created_at DESC
    """).fetchall()
    conn.close()
    return templates.TemplateResponse("admin_exam_sessions.html", {
        "request": request,
        "user": user,
        "classes": classes,
        "subjects": subjects,
        "test_types": test_types,
        "sessions": sessions,
    })


@router.post("/admin/exam-sessions/create")
async def admin_create_exam_session(request: Request):
    user = require_admin(request)
    form = await request.form()
    class_id = int(form["class_id"])
    test_type_id = int(form["test_type_id"])
    test_number = form["test_number"].strip()
    conduct_date = form["conduct_date"].strip()
    session_syllabus = form.get("session_syllabus", "").strip()

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO exam_sessions (class_id, test_type_id, test_number, conduct_date, syllabus)
        VALUES (?, ?, ?, ?, ?)
    """, (class_id, test_type_id, test_number, conduct_date, session_syllabus))
    session_id = cur.lastrowid

    idx = 0
    while True:
        subj_key = f"subject_id_{idx}"
        if subj_key not in form:
            break
        subject_id = int(form[subj_key])
        total_marks = float(form[f"total_marks_{idx}"])
        passing_marks = float(form[f"passing_marks_{idx}"])
        syllabus = form.get(f"syllabus_{idx}", "").strip()
        cur.execute("""
            INSERT INTO exam_session_subjects (session_id, subject_id, teacher_email, total_marks, passing_marks, syllabus)
            VALUES (?, ?, '', ?, ?, ?)
        """, (session_id, subject_id, total_marks, passing_marks, syllabus))
        idx += 1

    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-sessions?msg=created", status_code=303)


# ===================== ADMIN EXAM STATUS & CONFIRMATION =====================
@router.get("/admin/exam-status", response_class=HTMLResponse)
def admin_exam_status(request: Request):
    user = require_admin(request)
    conn = get_db()
    session_rows = conn.execute("""
        SELECT es.*, c.name as class_name, tt.name as test_type_name
        FROM exam_sessions es
        JOIN classes c ON c.id = es.class_id
        JOIN test_types tt ON tt.id = es.test_type_id
        ORDER BY es.conduct_date DESC
    """).fetchall()

    sessions = []
    for sess in session_rows:
        subjects = conn.execute("""
            SELECT ess.*, s.name as subject_name
            FROM exam_session_subjects ess
            JOIN subjects s ON s.id = ess.subject_id
            WHERE ess.session_id = ?
        """, (sess["id"],)).fetchall()
        sess_dict = dict(sess)
        sess_dict["subjects"] = [dict(s) for s in subjects]
        sessions.append(sess_dict)
    conn.close()
    return templates.TemplateResponse("admin_exam_status.html", {
        "request": request,
        "user": user,
        "sessions": sessions,
    })


def compute_and_store_summary(conn, session_id):
    session = conn.execute("SELECT * FROM exam_sessions WHERE id = ?", (session_id,)).fetchone()
    if not session:
        return
    class_id = session["class_id"]

    subjects = conn.execute("SELECT * FROM exam_session_subjects WHERE session_id = ?", (session_id,)).fetchall()
    if not subjects:
        return

    students = conn.execute("SELECT * FROM student_records WHERE class_id = ?", (class_id,)).fetchall()
    total_students = len(students)

    subject_details = {}
    overall_pass = 0
    overall_fail = 0

    for student in students:
        student_total_obtained = 0
        student_sum_passing = 0
        for subj in subjects:
            mark_row = conn.execute("""
                SELECT marks_obtained FROM exam_marks
                WHERE session_subject_id = ? AND student_id = ?
            """, (subj["id"], student["id"])).fetchone()
            obtained = mark_row["marks_obtained"] if mark_row and mark_row["marks_obtained"] is not None else 0
            total_marks = subj["total_marks"]
            passing_marks = subj["passing_marks"]

            student_total_obtained += obtained
            student_sum_passing += passing_marks

            key = str(subj["id"])
            if key not in subject_details:
                subject_details[key] = {
                    "subject_id": subj["subject_id"],
                    "subject_name": "",
                    "teacher_email": subj["teacher_email"],
                    "teacher_name": "",
                    "pass_count": 0,
                    "fail_count": 0,
                    "total_marks": total_marks,
                    "passing_marks": passing_marks
                }
                subj_name_row = conn.execute("SELECT name FROM subjects WHERE id = ?", (subj["subject_id"],)).fetchone()
                if subj_name_row:
                    subject_details[key]["subject_name"] = subj_name_row["name"]
                if subj["teacher_email"]:
                    teacher_row = conn.execute("SELECT full_name FROM users WHERE email = ?", (subj["teacher_email"],)).fetchone()
                    if teacher_row:
                        subject_details[key]["teacher_name"] = teacher_row["full_name"]

            if obtained >= passing_marks:
                subject_details[key]["pass_count"] += 1
            else:
                subject_details[key]["fail_count"] += 1

        if student_total_obtained >= student_sum_passing:
            overall_pass += 1
        else:
            overall_fail += 1

    subject_list = list(subject_details.values())

    conn.execute("DELETE FROM session_result_summary WHERE session_id = ?", (session_id,))
    conn.execute("""
        INSERT INTO session_result_summary (session_id, class_id, overall_pass_count, overall_fail_count, total_students, subject_details)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (session_id, class_id, overall_pass, overall_fail, total_students, json.dumps(subject_list)))


@router.post("/admin/exam-status/confirm/{session_id}")
def confirm_result(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("UPDATE exam_sessions SET status = 'confirmed' WHERE id = ?", (session_id,))
    compute_and_store_summary(conn, session_id)
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-status?msg=confirmed", status_code=303)


@router.post("/admin/exam-status/reopen/{session_id}")
def reopen_session(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("UPDATE exam_sessions SET status = 'open' WHERE id = ?", (session_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-status?msg=reopened", status_code=303)


# ===================== SESSION DELETE =====================
@router.post("/admin/exam-session/{session_id}/delete")
def delete_exam_session(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM exam_marks WHERE session_subject_id IN (SELECT id FROM exam_session_subjects WHERE session_id = ?)", (session_id,))
    conn.execute("DELETE FROM exam_session_subjects WHERE session_id = ?", (session_id,))
    conn.execute("DELETE FROM session_result_summary WHERE session_id = ?", (session_id,))
    conn.execute("DELETE FROM exam_sessions WHERE id = ?", (session_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-sessions?msg=deleted", status_code=303)


# ===================== TOGGLE VISIBILITY =====================
@router.post("/admin/exam-session/{session_id}/toggle-visibility")
def toggle_visibility(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    session = conn.execute("SELECT is_visible FROM exam_sessions WHERE id = ?", (session_id,)).fetchone()
    if session:
        new_val = 0 if session["is_visible"] else 1
        conn.execute("UPDATE exam_sessions SET is_visible = ? WHERE id = ?", (new_val, session_id))
        conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/exam-status?msg=visibility_updated", status_code=303)


# ===================== ADMIN EXAM CHARTS PAGE =====================
@router.get("/admin/exam-session/{session_id}/charts", response_class=HTMLResponse)
def admin_session_charts(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    session = conn.execute("""
        SELECT es.*, c.name as class_name
        FROM exam_sessions es JOIN classes c ON c.id = es.class_id
        WHERE es.id = ?
    """, (session_id,)).fetchone()
    if not session:
        conn.close()
        raise HTTPException(status_code=404, detail="Session not found")

    summary = conn.execute("SELECT * FROM session_result_summary WHERE session_id = ?", (session_id,)).fetchone()
    if not summary:
        conn.close()
        return RedirectResponse(url="/admin/exam-status?error=no_summary", status_code=303)

    subject_details = json.loads(summary["subject_details"])
    overall = {
        "pass": summary["overall_pass_count"],
        "fail": summary["overall_fail_count"],
        "total": summary["total_students"]
    }
    conn.close()
    return templates.TemplateResponse("admin_session_charts.html", {
        "request": request,
        "user": user,
        "session": dict(session),
        "overall": overall,
        "subjects": subject_details,
    })


# ===================== ADMIN TEST‑WISE RESULTS =====================
@router.get("/admin/test-results", response_class=HTMLResponse)
def admin_test_results(request: Request, test_type_id: int | None = None, test_number: str = ""):
    user = require_admin(request)
    conn = get_db()
    test_types = conn.execute("SELECT * FROM test_types ORDER BY name").fetchall()
    results = []
    if test_type_id:
        query = """
            SELECT es.id, es.class_id, es.test_number, c.name as class_name
            FROM exam_sessions es
            JOIN classes c ON c.id = es.class_id
            WHERE es.test_type_id = ?
        """
        params = [test_type_id]
        if test_number.strip():
            query += " AND es.test_number = ?"
            params.append(test_number.strip())
        sessions = conn.execute(query, params).fetchall()

        for sess in sessions:
            summary = conn.execute("SELECT * FROM session_result_summary WHERE session_id = ?", (sess["id"],)).fetchone()
            if summary:
                total = summary["total_students"]
                pass_count = summary["overall_pass_count"]
                fail_count = summary["overall_fail_count"]
                pass_percent = round((pass_count / total) * 100, 1) if total > 0 else 0
                results.append({
                    "class_name": sess["class_name"],
                    "session_id": sess["id"],
                    "test_number": sess["test_number"],
                    "total_students": total,
                    "pass_count": pass_count,
                    "fail_count": fail_count,
                    "pass_percent": pass_percent,
                })
    conn.close()
    return templates.TemplateResponse("admin_test_results.html", {
        "request": request,
        "user": user,
        "test_types": test_types,
        "selected_test_type_id": test_type_id,
        "selected_test_number": test_number,
        "results": results,
    })


# ===================== ADMIN EXPORT SESSION EXCEL (FINAL CLEAN) =====================
@router.get("/admin/exam-session/{session_id}/export")
def export_session_excel(session_id: int, request: Request):
    user = require_admin(request)
    conn = get_db()
    session = conn.execute("""
        SELECT es.*, c.name as class_name, tt.name as test_type_name
        FROM exam_sessions es
        JOIN classes c ON c.id = es.class_id
        JOIN test_types tt ON tt.id = es.test_type_id
        WHERE es.id = ?
    """, (session_id,)).fetchone()
    if not session:
        conn.close()
        raise HTTPException(status_code=404)

    subjects = conn.execute("""
        SELECT ess.*, s.name as subject_name
        FROM exam_session_subjects ess
        JOIN subjects s ON s.id = ess.subject_id
        WHERE ess.session_id = ?
    """, (session_id,)).fetchall()

    tpl = conn.execute("SELECT * FROM class_templates WHERE class_id = ?", (session["class_id"],)).fetchone()
    identity_cols = json.loads(tpl["identity_columns"]) if tpl else ["Student Name", "Father Name"]
    extra_cols = json.loads(tpl["extra_columns"]) if tpl else []

    students = conn.execute("""
        SELECT * FROM student_records WHERE class_id = ? ORDER BY name
    """, (session["class_id"],)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Row 1: School Name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1, value="Mustafa Public School")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.cell(row=1, column=1).alignment = center

    # Row 2: Session Info (compact)
    info_parts = []
    if session["class_name"]:
        info_parts.append(f"Class: {session['class_name']}")
    if session["test_type_name"]:
        info_parts.append(f"Test Type: {session['test_type_name']}")
    if session["test_number"]:
        info_parts.append(f"Test Number: {session['test_number']}")
    if session["conduct_date"]:
        info_parts.append(f"Date: {session['conduct_date']}")
    if session["syllabus"]:
        info_parts.append(f"Syllabus: {session['syllabus']}")

    info_text = "   |   ".join(info_parts)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(row=2, column=1, value=info_text)
    ws.cell(row=2, column=1).alignment = center

    # Header row
    header_row = 4
    all_header_cols = identity_cols + extra_cols
    col = 1
    for h in all_header_cols:
        ws.cell(row=header_row, column=col, value=h).font = bold
        ws.cell(row=header_row, column=col).alignment = center
        col += 1

    start_col = col
    for subj in subjects:
        ws.cell(row=header_row, column=col, value=f"{subj['subject_name']} (Out of {subj['total_marks']})").font = bold
        ws.cell(row=header_row, column=col).alignment = center
        col += 1

    total_marks_col = col
    ws.cell(row=header_row, column=col, value="Total Marks").font = bold
    ws.cell(row=header_row, column=col).alignment = center
    col += 1
    obtained_col = col
    ws.cell(row=header_row, column=col, value="Obtained").font = bold
    ws.cell(row=header_row, column=col).alignment = center
    col += 1
    percentage_col = col
    ws.cell(row=header_row, column=col, value="Percentage").font = bold
    ws.cell(row=header_row, column=col).alignment = center
    col += 1
    pass_fail_col = col
    ws.cell(row=header_row, column=col, value="Pass/Fail").font = bold
    ws.cell(row=header_row, column=col).alignment = center

    student_row = header_row + 1
    for student in students:
        col = 1
        ws.cell(row=student_row, column=col, value=student["name"]).alignment = center
        col += 1
        ws.cell(row=student_row, column=col, value=student["father_name"]).alignment = center
        col += 1
        extra_data = json.loads(student["extra_data"]) if student["extra_data"] else {}
        for ec in extra_cols:
            ws.cell(row=student_row, column=col, value=extra_data.get(ec, "")).alignment = center
            col += 1

        student_total_obtained = 0
        student_total_marks = 0
        student_sum_passing = 0
        for j, subj in enumerate(subjects):
            mark_row = conn.execute("""
                SELECT marks_obtained FROM exam_marks
                WHERE session_subject_id = ? AND student_id = ?
            """, (subj["id"], student["id"])).fetchone()
            marks = mark_row["marks_obtained"] if mark_row else 0
            if marks is None:
                marks = 0
            ws.cell(row=student_row, column=start_col + j, value=marks).alignment = center
            student_total_obtained += marks
            student_total_marks += subj["total_marks"]
            student_sum_passing += subj["passing_marks"]

        ws.cell(row=student_row, column=total_marks_col, value=student_total_marks).alignment = center
        ws.cell(row=student_row, column=obtained_col, value=student_total_obtained).alignment = center
        percentage = round((student_total_obtained / student_total_marks) * 100, 1) if student_total_marks else 0
        ws.cell(row=student_row, column=percentage_col, value=f"{percentage}%").alignment = center
        overall_pass = "Pass" if student_total_obtained >= student_sum_passing else "Fail"
        ws.cell(row=student_row, column=pass_fail_col, value=overall_pass).alignment = center
        if overall_pass == "Fail":
            ws.cell(row=student_row, column=pass_fail_col).font = Font(color="FF0000")
        student_row += 1

    # Teacher Summary (single table)
    teacher_start = student_row + 2
    ws.cell(row=teacher_start, column=1, value="Teacher Summary").font = bold
    teacher_start += 1

    teacher_headers = ["Teacher Name", "Subject", "Total Students", "Pass", "Fail", "Pass %"]
    for i, th in enumerate(teacher_headers, start=1):
        ws.cell(row=teacher_start, column=i, value=th).font = bold
        ws.cell(row=teacher_start, column=i).alignment = center
    teacher_start += 1

    for subj in subjects:
        teacher_email = subj["teacher_email"]
        if not teacher_email:
            continue
        teacher_name = ""
        teacher_row = conn.execute("SELECT full_name FROM users WHERE email = ?", (teacher_email,)).fetchone()
        if teacher_row:
            teacher_name = teacher_row["full_name"]
        marks_rows = conn.execute("""
            SELECT marks_obtained FROM exam_marks WHERE session_subject_id = ?
        """, (subj["id"],)).fetchall()
        pass_count = sum(1 for m in marks_rows if m["marks_obtained"] and m["marks_obtained"] >= subj["passing_marks"])
        fail_count = len(marks_rows) - pass_count
        total_students = len(marks_rows)
        pass_percent = round((pass_count / total_students) * 100, 1) if total_students else 0

        ws.cell(row=teacher_start, column=1, value=teacher_name).alignment = center
        ws.cell(row=teacher_start, column=2, value=subj["subject_name"]).alignment = center
        ws.cell(row=teacher_start, column=3, value=total_students).alignment = center
        ws.cell(row=teacher_start, column=4, value=pass_count).alignment = center
        ws.cell(row=teacher_start, column=5, value=fail_count).alignment = center
        ws.cell(row=teacher_start, column=6, value=f"{pass_percent}%").alignment = center
        teacher_start += 1

    # Class Summary
    summary = conn.execute("SELECT * FROM session_result_summary WHERE session_id = ?", (session_id,)).fetchone()
    if summary:
        total_students = summary["total_students"]
        overall_pass = summary["overall_pass_count"]
        overall_fail = summary["overall_fail_count"]
    else:
        total_students = len(students)
        overall_pass = 0
        for student in students:
            st_total_obt = 0
            st_sum_pass = 0
            for subj in subjects:
                mark = conn.execute("SELECT marks_obtained FROM exam_marks WHERE session_subject_id = ? AND student_id = ?",
                                    (subj["id"], student["id"])).fetchone()
                obt = mark["marks_obtained"] if mark else 0
                if obt is None:
                    obt = 0
                st_total_obt += obt
                st_sum_pass += subj["passing_marks"]
            if st_total_obt >= st_sum_pass:
                overall_pass += 1
        overall_fail = total_students - overall_pass

    pass_percent = round((overall_pass / total_students) * 100, 1) if total_students else 0
    class_start = teacher_start + 2
    ws.merge_cells(start_row=class_start, start_column=1, end_row=class_start, end_column=6)
    ws.cell(row=class_start, column=1, value="Class Summary").font = bold
    ws.cell(row=class_start, column=1).alignment = center
    class_start += 1
    labels = ["Total Students", "Pass", "Fail", "Pass %"]
    values = [total_students, overall_pass, overall_fail, f"{pass_percent}%"]
    for i, (label, val) in enumerate(zip(labels, values), start=1):
        ws.cell(row=class_start, column=i, value=label).font = bold
        ws.cell(row=class_start, column=i).alignment = center
        ws.cell(row=class_start + 1, column=i, value=val).alignment = center

    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"result_{session['class_name']}_{session['test_number']}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ===================== ADMIN SESSION SETTINGS =====================
@router.get("/admin/session-settings", response_class=HTMLResponse)
def session_settings_page(request: Request):
    user = require_admin(request)
    conn = get_db()
    start_date = get_setting(conn, "session_start_date", "")
    end_date = get_setting(conn, "session_end_date", "")
    conn.close()
    return templates.TemplateResponse("admin_session_settings.html", {
        "request": request,
        "user": user,
        "start_date": start_date,
        "end_date": end_date,
    })


@router.post("/admin/session-settings")
def save_session_settings(
    request: Request,
    session_start_date: str = Form(...),
    session_end_date: str = Form(...),
):
    user = require_admin(request)
    conn = get_db()
    set_setting(conn, "session_start_date", session_start_date.strip())
    set_setting(conn, "session_end_date", session_end_date.strip())
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin/session-settings?msg=saved", status_code=303)